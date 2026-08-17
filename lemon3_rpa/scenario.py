from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
import re

import yaml
from openpyxl import Workbook, load_workbook

ACTION_ALIASES = {
    "focus": "focus",
    "kich_app": "focus",
    "activate": "focus",
    "wait": "wait",
    "doi": "wait",
    "keys": "keys",
    "go_phim": "keys",
    "hotkey": "keys",
    "type": "type",
    "nhap": "type",
    "type_col": "type_col",
    "nhap_cot": "type_col",
    "enter": "enter",
    "tab": "tab",
    "alt": "alt",
    "click": "click",
    "click_xy": "click_xy",
    "click_toa_do": "click_xy",
    "pause": "pause",
    "tam_dung": "pause",
    "screenshot": "screenshot",
    "anh": "screenshot",
    "wait_window": "wait_window",
    "doi_cua_so": "wait_window",
    "focus_window": "focus_window",
    "kich_cua_so": "focus_window",
    "click_button": "click_button",
    "bam_nut": "click_button",
    "type_field": "type_field",
    "nhap_o": "type_field",
}


@dataclass
class Step:
    action: str
    value: str = ""
    wait_ms: int = 0
    note: str = ""
    extra: dict[str, Any] = field(default_factory=dict)


@dataclass
class Scenario:
    name: str
    title_contains: list[str]
    backend: str
    voucher_column: str
    skip_if_status: str
    delay_ms: int
    steps: list[Step]
    source: str
    flow: str = ""
    so_phieu_xy: str = ""
    # Kich thuoc cua so luc ghi toa do, de quy doi khi doi do phan giai.
    so_phieu_base: str = ""
    # Phim nong dung bot khi chuot dang bi bot gianh.
    stop_hotkey: str = "ctrl+shift+q"
    after_type_ms: int = 500
    # Cho toi da bao lau cho luoi loc xong. Thay phieu la di tiep ngay.
    filter_timeout_ms: int = 1000
    warehouse_code: str = "1000"
    warehouse_name: str = "Kho tổng JPT - Miền Nam"
    warehouse_xy: str = ""
    operation_code: str = "MNNKXB01"
    operation_xy: str = ""
    operation_row: int = 4
    # Nhip cho giua cac buoc xo/chon/xac nhan dropdown Loai nghiep vu.
    operation_ms: int = 300
    description_lines: list[dict[str, Any]] = field(default_factory=list)
    # 0 = tu dem so dong kho tren luoi. Dat 2/3 khi luoi khong doc duoc noi dung o.
    description_rows: int = 0
    description_label: str = "Dien giai"
    description_xy: str = ""
    # Rong = go xong de luoi tu loc. Dat '{ENTER}' hoac '{F5}' neu luoi doi phim.
    key_apply_filter: str = ""
    key_row_down: str = "{DOWN}"
    key_open_dropdown: str = "%{DOWN}"
    key_continue: str = "%t"
    key_goto_description: str = "{F11}"
    key_save: str = "%l"
    key_close: str = "%n"
    key_confirm: str = "{ENTER}"
    key_yes: str = "%y"


def default_scenario() -> Scenario:
    return Scenario(
        name="Xuat kho tu hoa don ban hang",
        title_contains=["DIGINET Desktop", "LEMON3-ERP", "D00F3000"],
        backend="win32",
        voucher_column="SỐ PHIẾU",
        skip_if_status="OK",
        delay_ms=400,
        steps=[],
        source="builtin",
        flow="xuat_kho",
        warehouse_code="1000",
        warehouse_name="Kho tổng JPT - Miền Nam",
        operation_code="MNNKXB01",
        description_lines=[
            {"template": "{so_phieu}", "required": True},
            {"template": "{DIEN_GIAI_2}", "required": False},
            {"template": "{DIEN_GIAI_3}", "required": False},
        ],
    )


def _as_int(value: Any, default: int = 0) -> int:
    try:
        if value is None or value == "":
            return default
        return int(float(value))
    except (TypeError, ValueError):
        return default


def load_scenario(path: str | Path | None) -> Scenario:
    if not path:
        return default_scenario()
    file_path = Path(path)
    if not file_path.exists():
        return default_scenario()
    suffix = file_path.suffix.lower()
    if suffix in {".yaml", ".yml"}:
        return _load_yaml(file_path)
    if suffix in {".xlsx", ".xlsm"}:
        return _load_excel(file_path)
    raise ValueError(f"Khong ho tro kich ban: {file_path.suffix}")


def _load_yaml(path: Path) -> Scenario:
    data = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    app = data.get("app") or {}
    loop = data.get("loop") or {}
    steps = []
    for raw in data.get("steps") or []:
        action = ACTION_ALIASES.get(str(raw.get("do") or raw.get("action") or "").strip().lower(), "")
        if not action:
            continue
        steps.append(
            Step(
                action=action,
                value=str(raw.get("keys") or raw.get("value") or raw.get("text") or raw.get("from") or ""),
                wait_ms=_as_int(raw.get("ms") or raw.get("wait_ms") or raw.get("cho_ms")),
                note=str(raw.get("note") or raw.get("ghi_chu") or ""),
                extra={k: v for k, v in raw.items() if k not in {"do", "action", "keys", "value", "text", "ms", "wait_ms", "cho_ms", "note", "ghi_chu"}},
            )
        )
    titles = app.get("title_contains") or app.get("title") or ["DIGINET Desktop", "LEMON3-ERP", "D00F3000"]
    if isinstance(titles, str):
        titles = [titles]
    erp = data.get("erp") or {}
    warehouse = erp.get("warehouse") or {}
    warehouse_code = erp.get("warehouse_code")
    warehouse_name = ""
    warehouse_xy = ""
    if isinstance(warehouse, dict):
        warehouse_code = warehouse_code or warehouse.get("code")
        warehouse_name = str(warehouse.get("name") or "")
        warehouse_xy = str(warehouse.get("xy") or "")
    elif isinstance(warehouse, str):
        warehouse_code = warehouse_code or warehouse
    desc = erp.get("description") or data.get("description") or {}
    lines = desc.get("lines") if isinstance(desc, dict) else desc
    description_lines = []
    for item in lines or []:
        if isinstance(item, str):
            description_lines.append({"template": item, "required": False})
        elif isinstance(item, dict):
            description_lines.append(
                {
                    "template": str(item.get("template") or item.get("value") or ""),
                    "required": bool(item.get("required", False)),
                }
            )
    if not description_lines:
        description_lines = default_scenario().description_lines
    operation = erp.get("operation") or {}
    if isinstance(operation, dict):
        operation_code = str(operation.get("code") or erp.get("operation_code") or "MNNKXB01")
        operation_xy = str(operation.get("xy") or "")
        operation_row = _as_int(operation.get("row"), 4)
        operation_ms = _as_int(operation.get("ms"), 300)
    else:
        operation_code = str(erp.get("operation_code") or operation or "MNNKXB01")
        operation_xy = ""
        operation_row = 4
        operation_ms = 300
    description_label = str(desc.get("label") or "Dien giai") if isinstance(desc, dict) else "Dien giai"
    description_xy = str(desc.get("xy") or "") if isinstance(desc, dict) else ""
    description_rows = _as_int(desc.get("rows"), 0) if isinstance(desc, dict) else 0
    keys = erp.get("keys") or {}
    return Scenario(
        name=str(data.get("name") or path.stem),
        title_contains=[str(t) for t in titles],
        backend=str(app.get("backend") or "win32"),
        voucher_column=str(loop.get("excel_column") or loop.get("voucher_column") or "SỐ PHIẾU"),
        skip_if_status=str(loop.get("skip_if_status") or "OK"),
        delay_ms=_as_int(app.get("delay_ms"), 400),
        steps=steps or default_scenario().steps,
        source=str(path),
        flow=str(data.get("flow") or ""),
        so_phieu_xy=str(app.get("so_phieu_xy") or ""),
        so_phieu_base=str(app.get("so_phieu_base") or ""),
        stop_hotkey=str(app.get("stop_hotkey") or "ctrl+shift+q"),
        after_type_ms=_as_int(app.get("after_type_ms"), 500),
        filter_timeout_ms=_as_int(app.get("filter_timeout_ms"), 1000),
        warehouse_code=str(warehouse_code or "1000"),
        warehouse_name=warehouse_name or "Kho tổng JPT - Miền Nam",
        warehouse_xy=warehouse_xy,
        operation_code=operation_code,
        operation_xy=operation_xy,
        operation_row=operation_row,
        operation_ms=operation_ms,
        description_lines=description_lines,
        description_rows=description_rows,
        description_label=description_label,
        description_xy=description_xy,
        key_apply_filter=str(keys.get("apply_filter") or ""),
        key_row_down=str(keys.get("row_down") or "{DOWN}"),
        key_open_dropdown=str(keys.get("dropdown") or "%{DOWN}"),
        key_continue=str(keys.get("continue") or "%t"),
        key_goto_description=str(keys.get("description") or "{F11}"),
        key_save=str(keys.get("save") or "%l"),
        key_close=str(keys.get("close") or "%n"),
        key_confirm=str(keys.get("ok") or "{ENTER}"),
        key_yes=str(keys.get("yes") or "%y"),
    )


def _load_excel(path: Path) -> Scenario:
    wb = load_workbook(path, data_only=True)
    sheet_name = "KichBan" if "KichBan" in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        scenario = default_scenario()
        scenario.source = str(path)
        return scenario
    headers = [_norm_header(c) for c in rows[0]]
    idx = {h: i for i, h in enumerate(headers) if h}
    steps: list[Step] = []
    for row in rows[1:]:
        if not row or all(v is None or str(v).strip() == "" for v in row):
            continue
        action_raw = _cell(row, idx, "hanhdong", "action", "do", "buoc_lam")
        action = ACTION_ALIASES.get(_norm_header(action_raw), "")
        if not action:
            continue
        steps.append(
            Step(
                action=action,
                value=str(_cell(row, idx, "giatri", "value", "keys", "text") or ""),
                wait_ms=_as_int(_cell(row, idx, "cho_ms", "wait_ms", "ms", "doi")),
                note=str(_cell(row, idx, "ghichu", "note", "ghi_chu") or ""),
            )
        )
    scenario = default_scenario()
    scenario.name = path.stem
    scenario.steps = steps or scenario.steps
    scenario.source = str(path)
    return scenario


def _norm_header(value: Any) -> str:
    return str(value or "").strip().lower().replace(" ", "").replace("_", "")


def _cell(row: tuple[Any, ...], idx: dict[str, int], *names: str) -> Any:
    for name in names:
        key = _norm_header(name)
        if key in idx:
            i = idx[key]
            if i < len(row):
                return row[i]
    return None


def save_yaml_template(path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        """name: Tim phieu Lemon3
app:
  title_contains:
    - LEMON3
    - Lemon3
    - DIGINET
  backend: uia          # thu uia; neu khong bat cua so thi doi win32
  delay_ms: 250
loop:
  excel_column: so_phieu
  skip_if_status: OK
steps:
  - do: focus
    note: Dua cua so Lemon3 len truoc
  - do: wait
    ms: 300
  - do: keys
    keys: "^f"
    note: Ctrl+F. Sua dung neu man hinh cua ban dung phim khac
  - do: wait
    ms: 200
  - do: keys
    keys: "^a"
    note: Chon het noi dung o tim
  - do: type_col
    from: so_phieu
  - do: wait
    ms: 150
  - do: enter
    ms: 800
    note: Sau khi tim thay phieu, them buoc thao tac ben duoi
  # - do: keys
  #   keys: "%c"
  #   note: Vi du Alt+C
  # - do: click_xy
  #   value: "120,80"
  #   note: Toa do trong cua so Lemon3. Dung nut Ghi toa do trong app
  # - do: pause
  #   note: Tam dung de kiem tra tay
""",
        encoding="utf-8",
    )


def save_excel_template(path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "KichBan"
    ws.append(["Buoc", "HanhDong", "GiaTri", "Cho_ms", "GhiChu"])
    ws.append([1, "focus", "", 300, "Dua Lemon3 len truoc"])
    ws.append([2, "keys", "^f", 200, "Ctrl+F mo tim kiem"])
    ws.append([3, "keys", "^a", 100, "Chon het o tim kiem"])
    ws.append([4, "type_col", "so_phieu", 150, "Go so phieu lay tu Excel"])
    ws.append([5, "enter", "", 800, "Nhan Enter de tim / mo phieu"])
    ws.append([6, "pause", "", 0, "Tam dung kiem tra. Xoa dong nay khi chay het tu dong"])
    for col, width in zip("ABCDE", (8, 14, 18, 10, 48)):
        ws.column_dimensions[col].width = width
    note = wb.create_sheet("HuongDan")
    note.append(["HanhDong"])
    note.append(["focus / kich_app", "Dua cua so Lemon3 len truoc"])
    note.append(["wait / doi", "Cho_ms = so miligiay"])
    note.append(["keys / go_phim", "GiaTri: ^f = Ctrl+F, %f = Alt+F, {ENTER}, {TAB}, {F3}"])
    note.append(["type / nhap", "Go chu. Dung {{so_phieu}} de lay cot Excel"])
    note.append(["type_col / nhap_cot", "GiaTri = ten cot Excel"])
    note.append(["enter", "Nhan Enter"])
    note.append(["tab", "Nhan Tab; GiaTri = so lan"])
    note.append(["click_xy / click_toa_do", "GiaTri = x,y tinh tu goc cua so Lemon3"])
    note.append(["pause / tam_dung", "Dung cho den khi bam Tiep tuc"])
    note.append(["screenshot / anh", "Luu anh man hinh vao logs/"])
    note.column_dimensions["A"].width = 28
    note.column_dimensions["B"].width = 70
    wb.save(path)
    wb.close()


def update_yaml_xy(path: str | Path, xy: str, base_size: str = "") -> None:
    file_path = Path(path)
    if file_path.suffix.lower() not in {".yaml", ".yml"} or not file_path.exists():
        return
    text = file_path.read_text(encoding="utf-8")
    text = _set_yaml_app_value(text, "so_phieu_xy", xy)
    if base_size:
        text = _set_yaml_app_value(text, "so_phieu_base", base_size)
    file_path.write_text(text, encoding="utf-8")


def _set_yaml_app_value(text: str, key: str, value: str) -> str:
    quoted = '"' + str(value).replace('"', "") + '"'
    if re.search(rf"^(\s*){key}\s*:", text, flags=re.M):
        return re.sub(
            rf"^(\s*){key}\s*:.*$",
            rf"\g<1>{key}: {quoted}",
            text,
            count=1,
            flags=re.M,
        )
    return re.sub(
        r"^(app:\s*)$",
        rf"\g<1>\n  {key}: {quoted}",
        text,
        count=1,
        flags=re.M,
    )


def update_yaml_automation(path: str | Path, values: dict[str, Any]) -> None:
    """Cap nhat cac tham so automation, giu nguyen cac muc YAML khac."""
    file_path = Path(path)
    if file_path.suffix.lower() not in {".yaml", ".yml"}:
        raise ValueError("Cai dat automation chi luu duoc vao file YAML.")
    if not file_path.exists():
        raise FileNotFoundError(f"Khong tim thay kich ban: {file_path}")

    data = yaml.safe_load(file_path.read_text(encoding="utf-8")) or {}
    app = data.setdefault("app", {})
    erp = data.setdefault("erp", {})
    warehouse = erp.setdefault("warehouse", {})
    operation = erp.setdefault("operation", {})
    keys = erp.setdefault("keys", {})

    app["delay_ms"] = int(values["delay_ms"])
    app["after_type_ms"] = int(values["after_type_ms"])
    app["stop_hotkey"] = str(values["stop_hotkey"]).strip().lower()
    warehouse["code"] = str(values["warehouse_code"]).strip()
    warehouse["name"] = str(values["warehouse_name"]).strip()
    operation["code"] = str(values["operation_code"]).strip()
    operation["row"] = int(values["operation_row"])
    operation["ms"] = int(values["operation_ms"])

    key_map = {
        "row_down": "key_row_down",
        "dropdown": "key_open_dropdown",
        "continue": "key_continue",
        "description": "key_goto_description",
        "save": "key_save",
        "close": "key_close",
        "yes": "key_yes",
        "ok": "key_confirm",
    }
    for yaml_key, value_key in key_map.items():
        keys[yaml_key] = str(values[value_key]).strip()

    file_path.write_text(
        yaml.safe_dump(data, allow_unicode=True, sort_keys=False, width=120),
        encoding="utf-8",
    )
