from __future__ import annotations

import atexit
import re
import time
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
import csv

VOUCHER_ALIASES = (
    "so_phieu",
    "sophieu",
    "so phieu",
    "số phiếu",
    "mã phiếu",
    "ma_phieu",
    "maphieu",
    "voucher",
    "voucher_no",
    "docno",
    "soct",
    "so_ct",
    "số ct",
)

STATUS_ALIASES = ("ket_qua", "ketqua", "kết quả", "status", "trang_thai")


def _norm(text: Any) -> str:
    return re.sub(r"\s+", " ", str(text or "").strip().lower())


def _header_map(headers: list[str]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, raw in enumerate(headers):
        key = _norm(raw)
        if key:
            mapping[key] = idx
            mapping[str(raw).strip()] = idx
    return mapping


COL_LETTER_RE = re.compile(r"^[A-Za-z]{1,2}$")


def _col_letter_to_index(letter: str) -> int:
    n = 0
    for ch in letter.strip().upper():
        n = n * 26 + (ord(ch) - 64)
    return n - 1


def find_voucher_column(headers: list[str], preferred: str | None = None) -> str | None:
    mapped = _header_map(headers)
    if preferred:
        raw = str(preferred).strip()
        if COL_LETTER_RE.match(raw):
            idx = _col_letter_to_index(raw)
            if 0 <= idx < len(headers):
                return headers[idx]
        key = _norm(raw)
        if key in mapped:
            return headers[mapped[key]]
        for h in headers:
            if _norm(h) == key:
                return h
    for alias in VOUCHER_ALIASES:
        if alias in mapped:
            return headers[mapped[alias]]
    if headers:
        return headers[0]
    return None


def load_rows(path: str | Path, sheet: str | None = None) -> tuple[list[str], list[dict[str, Any]]]:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        return _load_csv(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        wb.close()
        return [], []
    headers = [str(c).strip() if c is not None else f"cot_{i+1}" for i, c in enumerate(header_row)]
    data: list[dict[str, Any]] = []
    for excel_row, values in enumerate(rows_iter, start=2):
        if values is None or all(v is None or str(v).strip() == "" for v in values):
            continue
        item: dict[str, Any] = {"_excel_row": excel_row}
        for i, header in enumerate(headers):
            value = values[i] if i < len(values) else None
            item[header] = value
            item[_norm(header)] = value
        data.append(item)
    wb.close()
    return headers, data


def _load_csv(path: Path) -> tuple[list[str], list[dict[str, Any]]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.reader(handle)
        try:
            header_row = next(reader)
        except StopIteration:
            return [], []
        headers = [str(c).strip() if c else f"cot_{i+1}" for i, c in enumerate(header_row)]
        data: list[dict[str, Any]] = []
        for excel_row, values in enumerate(reader, start=2):
            if not values or all(str(v).strip() == "" for v in values):
                continue
            item: dict[str, Any] = {"_excel_row": excel_row}
            for i, header in enumerate(headers):
                value = values[i] if i < len(values) else None
                item[header] = value
                item[_norm(header)] = value
            data.append(item)
        return headers, data


def write_status(path: str | Path, row_number: int, status: str, message: str = "", sheet: str | None = None) -> None:
    path = Path(path)
    if path.suffix.lower() == ".csv":
        _write_status_csv(path, row_number, status, message)
        return
    wb = load_workbook(path)
    ws = wb[sheet] if sheet and sheet in wb.sheetnames else wb[wb.sheetnames[0]]
    headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
    mapped = _header_map(headers)

    def ensure_col(name: str, aliases: tuple[str, ...]) -> int:
        for alias in aliases:
            if alias in mapped:
                return mapped[alias] + 1
        col = len(headers) + 1
        ws.cell(1, col, name)
        headers.append(name)
        mapped[_norm(name)] = col - 1
        return col

    status_col = ensure_col("KetQua", STATUS_ALIASES)
    msg_col = ensure_col("GhiChu_RPA", ("ghichu_rpa", "ghi chú rpa", "error", "loi"))
    ws.cell(row_number, status_col, status)
    ws.cell(row_number, msg_col, message[:32000])
    wb.save(path)
    wb.close()


def _write_status_csv(path: Path, row_number: int, status: str, message: str) -> None:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        return
    headers = rows[0]
    mapped = {_norm(h): i for i, h in enumerate(headers)}

    def ensure(name: str, aliases: tuple[str, ...]) -> int:
        for alias in aliases:
            if alias in mapped:
                return mapped[alias]
        headers.append(name)
        mapped[_norm(name)] = len(headers) - 1
        for row in rows[1:]:
            row.append("")
        return len(headers) - 1

    status_i = ensure("KetQua", STATUS_ALIASES)
    msg_i = ensure("GhiChu_RPA", ("ghichu_rpa", "ghi chú rpa", "error", "loi"))
    idx = row_number - 1
    if 0 < idx < len(rows):
        while len(rows[idx]) < len(headers):
            rows[idx].append("")
        rows[idx][status_i] = status
        rows[idx][msg_i] = message[:32000]
    rows[0] = headers
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        csv.writer(handle).writerows(rows)


def _write_extras_csv(path: Path, row_number: int, extras: dict[str, Any]) -> None:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.reader(handle))
    if not rows:
        return
    headers = rows[0]
    mapped = {_norm(h): i for i, h in enumerate(headers)}
    for name in extras:
        if _norm(name) not in mapped:
            headers.append(name)
            mapped[_norm(name)] = len(headers) - 1
            for row in rows[1:]:
                row.append("")
    idx = row_number - 1
    if 0 < idx < len(rows):
        while len(rows[idx]) < len(headers):
            rows[idx].append("")
        for name, value in extras.items():
            rows[idx][mapped[_norm(name)]] = str(value)
    rows[0] = headers
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        csv.writer(handle).writerows(rows)


def copy_input(source: str | Path, dest: str | Path) -> Path:
    import shutil

    source = Path(source)
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, dest)
    return dest


# File ket qua giu dinh dang cua Excel goc nen load+save ton ~10s. Giu trong bo nho,
# ghi xuong dia theo chu ky de moi phieu khong phai cho.
SAVE_INTERVAL_S = 15.0
_BOOKS: dict[str, dict[str, Any]] = {}


def _book(path: Path) -> dict[str, Any]:
    key = str(path.resolve())
    entry = _BOOKS.get(key)
    if entry is None:
        wb = load_workbook(path)
        ws = wb[wb.sheetnames[0]]
        headers = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        entry = {
            "wb": wb,
            "ws": ws,
            "path": path,
            "headers": headers,
            "mapped": _header_map(headers),
            "dirty": False,
            "saved_at": time.monotonic(),
        }
        _BOOKS[key] = entry
    return entry


def _save_book(entry: dict[str, Any]) -> None:
    entry["wb"].save(entry["path"])
    entry["dirty"] = False
    entry["saved_at"] = time.monotonic()


def flush_results(path: str | Path | None = None) -> None:
    """Ghi phan ket qua con nam trong bo nho xuong file."""
    keys = [str(Path(path).resolve())] if path else list(_BOOKS)
    for key in keys:
        entry = _BOOKS.get(key)
        if entry and entry["dirty"]:
            try:
                _save_book(entry)
            except Exception:
                pass


def close_results(path: str | Path | None = None) -> None:
    flush_results(path)
    keys = [str(Path(path).resolve())] if path else list(_BOOKS)
    for key in keys:
        entry = _BOOKS.pop(key, None)
        if entry:
            try:
                entry["wb"].close()
            except Exception:
                pass


atexit.register(flush_results)


def write_run_result(
    path: str | Path,
    row_number: int,
    status: str,
    message: str = "",
    attempts: int = 1,
    run_id: str = "",
    screenshot_path: str = "",
    erp_reference: str = "",
) -> None:
    from datetime import datetime

    path = Path(path)
    extras = {
        "AUTOMATION_STATUS": status,
        "AUTOMATION_MESSAGE": message[:32000],
        "ATTEMPTS": attempts,
        "RUN_ID": run_id,
        "PROCESSED_AT": datetime.now().isoformat(timespec="seconds"),
        "SCREENSHOT_PATH": screenshot_path,
        "ERP_REFERENCE": erp_reference,
    }
    if path.suffix.lower() == ".csv":
        _write_extras_csv(path, row_number, extras)
        return
    entry = _book(path)
    ws = entry["ws"]
    headers = entry["headers"]
    mapped = entry["mapped"]

    def col_for(name: str) -> int:
        key = _norm(name)
        if key in mapped:
            return mapped[key] + 1
        idx = len(headers) + 1
        ws.cell(1, idx, name)
        headers.append(name)
        mapped[key] = idx - 1
        return idx

    for name, value in extras.items():
        ws.cell(row_number, col_for(name), value)
    entry["dirty"] = True
    if time.monotonic() - entry["saved_at"] >= SAVE_INTERVAL_S:
        _save_book(entry)


def create_sample_workbook(path: str | Path) -> None:
    path = Path(path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Phieu"
    ws.append(["so_phieu", "ghi_chu", "KetQua", "GhiChu_RPA"])
    ws.append(["HD001", "Vi du 1", "", ""])
    ws.append(["HD002", "Vi du 2", "", ""])
    ws.append(["HD003", "Vi du 3", "", ""])
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 40
    wb.save(path)
    wb.close()
